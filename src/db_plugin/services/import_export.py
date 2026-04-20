import csv
import json
import logging
from pathlib import Path

import openpyxl

from db_plugin.core.executor import QueryExecutor
from db_plugin.models.result import QueryResult

logger = logging.getLogger(__name__)

BATCH_SIZE = 100


class ImportExportService:
    """Import and export data to/from CSV, Excel, and JSON."""

    def __init__(self, executor: QueryExecutor):
        self.executor = executor

    def _quote_col(self, dialect, name: str) -> str:
        return dialect.quote_identifier(name)

    def _batch_insert(self, dialect, table: str, records: list[dict]) -> tuple[int, int]:
        """Insert records in batches. Returns (inserted, errors)."""
        inserted = 0
        errors = 0
        for batch_start in range(0, len(records), BATCH_SIZE):
            batch = records[batch_start:batch_start + BATCH_SIZE]
            placeholders = []
            all_values = []
            cols = list(batch[0].keys())
            for record in batch:
                placeholders.append(f"({', '.join(['%s'] * len(cols))})")
                all_values.extend(record.get(c) for c in cols)
            col_list = ", ".join(self._quote_col(dialect, c) for c in cols)
            sql = f"INSERT INTO {dialect.format_table_ref(table)} ({col_list}) VALUES {', '.join(placeholders)}"
            try:
                result = dialect.execute_query(sql, tuple(all_values))
                if result.error_message:
                    for record in batch:
                        single = dialect.insert(table, record)
                        if single.error_message is None:
                            inserted += 1
                        else:
                            errors += 1
                            logger.warning("Fallback insert failed into '%s': %s", table, single.error_message)
                else:
                    inserted += len(batch)
            except Exception as e:
                errors += len(batch)
                logger.warning("Batch insert failed for '%s': %s, falling back to individual", table, e)
                for record in batch:
                    single = dialect.insert(table, record)
                    if single.error_message is None:
                        inserted += 1
                    else:
                        errors += 1
                        logger.warning("Fallback insert failed into '%s': %s", table, single.error_message)
        return inserted, errors

    def export_csv(self, result: QueryResult, filepath: str) -> None:
        logger.info("Exporting %d rows to CSV: %s", result.row_count, filepath)
        path = Path(filepath)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=result.columns)
            writer.writeheader()
            for row in result.rows:
                writer.writerow(row)
        logger.info("CSV export complete")

    def export_excel(self, result: QueryResult, filepath: str) -> None:
        logger.info("Exporting %d rows to Excel: %s", result.row_count, filepath)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Query Result"
        ws.append(result.columns)
        for row in result.rows:
            ws.append([row.get(col) for col in result.columns])
        wb.save(filepath)
        logger.info("Excel export complete")

    def export_json(self, result: QueryResult, filepath: str) -> None:
        logger.info("Exporting %d rows to JSON: %s", result.row_count, filepath)
        path = Path(filepath)
        path.write_text(json.dumps(result.rows, indent=2, default=str), encoding="utf-8")
        logger.info("JSON export complete")

    def import_csv(self, filepath: str, table: str) -> int:
        logger.info("Importing CSV %s into table '%s'", filepath, table)
        dialect = self.executor.connection.get_dialect()
        path = Path(filepath)
        records = []
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                records.append(dict(row))
        if not records:
            return 0
        inserted, errors = self._batch_insert(dialect, table, records)
        logger.info("CSV import complete: %d rows inserted, %d errors into '%s'", inserted, errors, table)
        return inserted

    def import_excel(self, filepath: str, table: str) -> int:
        logger.info("Importing Excel %s into table '%s'", filepath, table)
        dialect = self.executor.connection.get_dialect()
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            records.append(data)
        if not records:
            return 0
        inserted, errors = self._batch_insert(dialect, table, records)
        logger.info("Excel import complete: %d rows inserted, %d errors into '%s'", inserted, errors, table)
        return inserted

    def import_json(self, filepath: str, table: str) -> int:
        logger.info("Importing JSON %s into table '%s'", filepath, table)
        dialect = self.executor.connection.get_dialect()
        path = Path(filepath)
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            data = [data]
        if not data:
            return 0
        inserted, errors = self._batch_insert(dialect, table, data)
        logger.info("JSON import complete: %d rows inserted, %d errors into '%s'", inserted, errors, table)
        return inserted
